from openpyxl import load_workbook

def extract_text(file_path):

    wb = load_workbook(file_path)

    text = []

    for sheet in wb:

        for row in sheet.iter_rows():

            values = []

            for cell in row:

                values.append(
                    str(cell.value)
                )

            text.append(
                " ".join(values)
            )

    return "\n".join(text)